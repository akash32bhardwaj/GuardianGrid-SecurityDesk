"""
resident_db.py — GuardianGrid Resident Vehicle Database
---------------------------------------------------------
Manages resident vehicle data using PostgreSQL for Render deployment.
"""

import csv
import logging
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional
from db import _conn

logger = logging.getLogger(__name__)

# ── Try importing openpyxl for Excel support ─────────────────────
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Resident dataclass ───────────────────────────────────────────
@dataclass
class Resident:
    plate_number:   str
    resident_name:  str
    flat_number:    str
    block:          str        = ""
    phone:          str        = ""
    vehicle_type:   str        = "Car"
    vehicle_model:  str        = ""
    vehicle_color:  str        = ""
    status:         str        = "KNOWN"   # KNOWN / BLACKLISTED / VISITOR
    notes:          str        = ""
    added_on:       str        = field(
        default_factory=lambda: datetime.now().strftime("%Y-%m-%d")
    )

    def to_dict(self):
        return asdict(self)

    @property
    def display_name(self):
        parts = [self.resident_name]
        if self.flat_number:
            parts.append(f"Flat {self.flat_number}")
        if self.block:
            parts.append(f"Block {self.block}")
        return " · ".join(parts)

_RESIDENT_SCHEMA = """
CREATE TABLE IF NOT EXISTS resident_profiles (
    plate_number   VARCHAR(50) PRIMARY KEY,
    resident_name  VARCHAR(255),
    flat_number    VARCHAR(50),
    block          VARCHAR(50),
    phone          VARCHAR(50),
    vehicle_type   VARCHAR(50),
    vehicle_model  VARCHAR(100),
    vehicle_color  VARCHAR(50),
    status         VARCHAR(50),
    notes          TEXT,
    added_on       VARCHAR(50)
);
"""

# ── Database class ───────────────────────────────────────────────
class ResidentDatabase:
    def __init__(self):
        try:
            with _conn() as c:
                with c.cursor() as cur:
                    cur.execute(_RESIDENT_SCHEMA)
            logger.info("Resident profiles table ready.")
        except Exception as e:
            logger.error(f"Failed to initialize resident DB: {e}")

    def _row_to_resident(self, row):
        return Resident(
            plate_number=row[0],
            resident_name=row[1],
            flat_number=row[2],
            block=row[3],
            phone=row[4],
            vehicle_type=row[5],
            vehicle_model=row[6],
            vehicle_color=row[7],
            status=row[8],
            notes=row[9],
            added_on=row[10]
        )

    # ── Core operations ───────────────────────────────────────
    def lookup(self, plate: str) -> Optional[Resident]:
        clean = plate.upper().replace(" ", "")
        with _conn() as c:
            with c.cursor() as cur:
                # Exact match
                cur.execute("SELECT * FROM resident_profiles WHERE plate_number = %s", (clean,))
                row = cur.fetchone()
                if row:
                    return self._row_to_resident(row)
                
                # Partial match
                cur.execute("SELECT * FROM resident_profiles")
                rows = cur.fetchall()
                for r in rows:
                    stored_plate = r[0]
                    if clean in stored_plate or stored_plate in clean:
                        return self._row_to_resident(r)
        return None

    def fuzzy_lookup(self, plate: str, max_distance: int = 1) -> Optional[Resident]:
        clean = plate.upper().replace(" ", "")
        with _conn() as c:
            with c.cursor() as cur:
                cur.execute("SELECT * FROM resident_profiles WHERE plate_number = %s", (clean,))
                row = cur.fetchone()
                if row:
                    return self._row_to_resident(row)

                cur.execute("SELECT * FROM resident_profiles")
                rows = cur.fetchall()
                
        candidates = []
        for r in rows:
            stored_plate = r[0]
            if len(stored_plate) != len(clean):
                continue
            dist = sum(1 for a, b in zip(clean, stored_plate) if a != b)
            if dist <= max_distance:
                candidates.append((dist, self._row_to_resident(r)))

        if len(candidates) == 1:
            return candidates[0][1]
        if len(candidates) > 1:
            candidates.sort(key=lambda c: c[0])
            if candidates[0][0] < candidates[1][0]:
                return candidates[0][1]
        return None

    def add(self, resident: Resident):
        clean = resident.plate_number.upper().replace(" ", "")
        resident.plate_number = clean
        with _conn() as c:
            with c.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO resident_profiles 
                    (plate_number, resident_name, flat_number, block, phone, vehicle_type, vehicle_model, vehicle_color, status, notes, added_on)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (plate_number) DO UPDATE SET
                        resident_name = EXCLUDED.resident_name,
                        flat_number = EXCLUDED.flat_number,
                        block = EXCLUDED.block,
                        phone = EXCLUDED.phone,
                        vehicle_type = EXCLUDED.vehicle_type,
                        vehicle_model = EXCLUDED.vehicle_model,
                        vehicle_color = EXCLUDED.vehicle_color,
                        status = EXCLUDED.status,
                        notes = EXCLUDED.notes,
                        added_on = EXCLUDED.added_on
                    """,
                    (resident.plate_number, resident.resident_name, resident.flat_number, resident.block, resident.phone, resident.vehicle_type, resident.vehicle_model, resident.vehicle_color, resident.status, resident.notes, resident.added_on)
                )

    def remove(self, plate: str) -> bool:
        clean = plate.upper().replace(" ", "")
        with _conn() as c:
            with c.cursor() as cur:
                cur.execute("DELETE FROM resident_profiles WHERE plate_number = %s", (clean,))
                return cur.rowcount > 0

    def get_all(self) -> list:
        with _conn() as c:
            with c.cursor() as cur:
                cur.execute("SELECT * FROM resident_profiles")
                rows = cur.fetchall()
        return [self._row_to_resident(r).to_dict() for r in rows]

    def count(self) -> int:
        with _conn() as c:
            with c.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM resident_profiles")
                res = cur.fetchone()
                return res[0] if res else 0

    def blacklist(self, plate: str, reason: str = ""):
        r = self.lookup(plate)
        if r:
            r.status = "BLACKLISTED"
            r.notes  = reason
            self.add(r)

    # ── Excel import ──────────────────────────────────────────
    def import_from_excel(self, filepath: str) -> dict:
        if not EXCEL_AVAILABLE:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(filepath)
        if not path.exists():
            return {"error": f"File not found: {filepath}"}

        try:
            wb   = openpyxl.load_workbook(path)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return {"error": f"Cannot read Excel file: {e}"}

        if not rows:
            return {"error": "Excel file is empty"}

        headers = [str(h).lower().strip() if h else "" for h in rows[0]]

        col_map = {
            "plate_number":  ["plate", "plate_number", "vehicle_number", "registration", "reg_no", "number_plate"],
            "resident_name": ["name", "resident_name", "owner", "resident", "full_name", "owner_name"],
            "flat_number":   ["flat", "flat_number", "flat_no", "unit", "apartment", "house_no", "unit_no"],
            "block":         ["block", "tower", "wing", "sector"],
            "phone":         ["phone", "mobile", "contact", "phone_number", "mobile_number"],
            "vehicle_type":  ["type", "vehicle_type", "vehicle"],
            "vehicle_model": ["model", "vehicle_model", "car_model"],
            "vehicle_color": ["color", "colour", "vehicle_color"],
            "notes":         ["notes", "remarks", "comment"],
        }

        col_idx = {}
        for field_name, aliases in col_map.items():
            for alias in aliases:
                if alias in headers:
                    col_idx[field_name] = headers.index(alias)
                    break

        if "plate_number" not in col_idx:
            return {"error": "Could not find plate number column. Please name it 'plate' or 'plate_number'."}

        imported = 0
        skipped  = 0
        errors   = []

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                plate = str(row[col_idx["plate_number"]] or "").strip()
                if not plate or plate.lower() == "none":
                    skipped += 1
                    continue

                def get_col(field, default=""):
                    idx = col_idx.get(field)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val is not None else default
                    return default

                resident = Resident(
                    plate_number  = plate.upper().replace(" ", ""),
                    resident_name = get_col("resident_name", "Unknown"),
                    flat_number   = get_col("flat_number"),
                    block         = get_col("block"),
                    phone         = get_col("phone"),
                    vehicle_type  = get_col("vehicle_type", "Car"),
                    vehicle_model = get_col("vehicle_model"),
                    vehicle_color = get_col("vehicle_color"),
                    notes         = get_col("notes"),
                    status        = "KNOWN",
                )
                self.add(resident)
                imported += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {e}")
                skipped += 1

        return {
            "imported": imported,
            "skipped":  skipped,
            "total":    len(rows) - 1,
            "errors":   errors,
        }

    # ── Excel export ──────────────────────────────────────────
    def export_to_excel(self, filepath: str = "data/residents_export.xlsx"):
        if not EXCEL_AVAILABLE:
            return {"error": "openpyxl not installed"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Residents"

        headers = [
            "Plate Number", "Resident Name", "Flat Number", "Block",
            "Phone", "Vehicle Type", "Vehicle Model",
            "Vehicle Color", "Status", "Notes", "Added On"
        ]
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")

        with _conn() as c:
            with c.cursor() as cur:
                cur.execute("SELECT * FROM resident_profiles")
                rows = cur.fetchall()

        for row in rows:
            r = self._row_to_resident(row)
            ws.append([
                r.plate_number, r.resident_name, r.flat_number,
                r.block, r.phone, r.vehicle_type, r.vehicle_model,
                r.vehicle_color, r.status, r.notes, r.added_on
            ])

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        wb.save(filepath)
        return {"exported": len(rows), "file": filepath}

    # ── CSV import (fallback if no Excel) ─────────────────────
    def import_from_csv(self, filepath: str) -> dict:
        path = Path(filepath)
        if not path.exists():
            return {"error": f"File not found: {filepath}"}

        imported = 0
        skipped  = 0

        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                plate = (row.get("plate_number") or
                         row.get("plate") or "").strip().upper()
                if not plate:
                    skipped += 1
                    continue
                resident = Resident(
                    plate_number  = plate.replace(" ", ""),
                    resident_name = row.get("resident_name", "Unknown"),
                    flat_number   = row.get("flat_number", ""),
                    block         = row.get("block", ""),
                    phone         = row.get("phone", ""),
                    vehicle_type  = row.get("vehicle_type", "Car"),
                    vehicle_model = row.get("vehicle_model", ""),
                    vehicle_color = row.get("vehicle_color", ""),
                    notes         = row.get("notes", ""),
                )
                self.add(resident)
                imported += 1

        return {"imported": imported, "skipped": skipped}


# ── Global instance ───────────────────────────────────────────────
db = ResidentDatabase()

# ── Standalone test ───────────────────────────────────────────────
if __name__ == "__main__":
    print("Testing Resident Database with PostgreSQL...")
    db.add(Resident(
        plate_number="PB08EY5332", resident_name="Akash Singh",
        flat_number="302", block="B", phone="98765-43210",
        vehicle_type="Car", vehicle_model="Hyundai i20",
        vehicle_color="White"
    ))
    db.add(Resident(
        plate_number="PB10AB2025", resident_name="Gurpreet Kaur",
        flat_number="105", block="A", phone="98100-12345",
        vehicle_type="Car", vehicle_model="Maruti Swift",
        vehicle_color="Silver"
    ))
    db.add(Resident(
        plate_number="DL3CAB5678", resident_name="Rajesh Kumar",
        flat_number="201", block="C", phone="99999-88888",
        vehicle_type="Car", vehicle_model="Honda City",
        vehicle_color="Black", status="BLACKLISTED",
        notes="Unpaid dues"
    ))

    print(f"\nTotal residents: {db.count()}")

    result = db.lookup("PB08EY5332")
    if result:
        print(f"\nLookup PB08EY5332:")
        print(f"  Name  : {result.resident_name}")
        print(f"  Flat  : {result.display_name}")
        print(f"  Phone : {result.phone}")
        print(f"  Car   : {result.vehicle_color} {result.vehicle_model}")
        print(f"  Status: {result.status}")

    print("\nDatabase test complete! ✅")
