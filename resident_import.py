r"""
resident_import.py — DEFENDER OCTA bulk onboarding (import · summary · invite)
------------------------------------------------------------------------------
"Hand me your society's Excel sheet and it's live by Friday."

  POST /api/admin/residents/import      multipart: file=<xlsx|csv>, dry_run=1|0
                                        -> {flats, vehicles, issues[], preview[]}
                                        dry_run=1 (default) checks only; 0 writes
  GET  /api/admin/residents/summary     counts for the Residents page
  GET  /api/admin/residents/template    download the Excel template
  POST /api/admin/residents/invite      {phones?: [...], sample?: N, message?}
                                        -> WhatsApp welcome to every mobile on
                                        file (or just `phones`, or the first N)
  GET  /api/admin/residents/invite/status

All routes sit behind the dashboard JWT (global guard); VIEWER role is
already refused on non-GET by api_server.

SHEET FORMAT (headers are matched loosely — case/spaces/underscores ignored)
  block | flat | owner name | mobile | vehicle 1 | vehicle 2 | vehicle 3 | model
  - one row per flat, vehicles across columns   (recommended, see template)
  - OR one row per vehicle with the flat repeated (also accepted)
  - "flat" may already contain the block ("B-302") — then block is optional

WHERE IT WRITES
  vehicles  -> resident_db (residents.json)   via resident_db.db.add()
  flats     -> flat_directory                 via its own upsert function if it
               has one (upsert_flat / add_flat / set_flat / save_flat), else
               directly into its table (auto-detected columns).

Integration in api_server.py:
    from resident_import import resident_import_bp, init_resident_import
    init_resident_import(base_dir=BASE_DIR)
    app.register_blueprint(resident_import_bp)
"""

import csv
import io
import logging
import os
import re
import sqlite3
import threading
import time
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

logger = logging.getLogger(__name__)
resident_import_bp = Blueprint("resident_import", __name__)

_BASE_DIR = "."
_DB_PATH = "guardiangrid.db"
_invite_job = {"running": False, "total": 0, "sent": 0, "failed": 0,
               "started": None, "finished": None, "last_error": ""}
_lock = threading.Lock()

PLATE_RE = re.compile(r"^[A-Z]{2}\d{1,2}[A-Z]{0,3}\d{3,4}$")


def init_resident_import(base_dir: str):
    global _BASE_DIR, _DB_PATH
    _BASE_DIR = base_dir
    docker_db = "/data/guardiangrid.db"
    _DB_PATH = docker_db if os.path.exists(docker_db) \
        else os.path.join(base_dir, "guardiangrid.db")
    try:
        con = sqlite3.connect(_DB_PATH)
        con.executescript("""
        CREATE TABLE IF NOT EXISTS resident_invites (
            phone TEXT PRIMARY KEY, flat_no TEXT, sent_at TEXT, status TEXT, detail TEXT
        );
        CREATE TABLE IF NOT EXISTS resident_imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT, flats INTEGER,
            vehicles INTEGER, issues INTEGER, imported_at TEXT, by_user TEXT
        );""")
        con.commit(); con.close()
    except sqlite3.Error as e:
        logger.error(f"[IMPORT] table init: {e}")
    logger.info("[IMPORT] resident importer armed")


# ════════════════════════════════════════════════════════════════════
# Normalisers
# ════════════════════════════════════════════════════════════════════

def _digits(s): return re.sub(r"\D", "", str(s or ""))


def _norm_phone(s):
    d = _digits(s)
    if len(d) == 10: return "+91" + d
    if len(d) == 12 and d.startswith("91"): return "+" + d
    if len(d) == 11 and d.startswith("0"): return "+91" + d[1:]
    return ""


def _norm_plate(p): return re.sub(r"[^A-Z0-9]", "", str(p or "").upper())


def _norm_flat(block, flat):
    f = re.sub(r"\s+", "", str(flat or "").upper().strip())
    b = re.sub(r"\s+", "", str(block or "").upper().strip())
    if not f:
        return ""
    if b and not f.startswith(b + "-") and not f.startswith(b):
        return f"{b}-{f}"
    if b and f.startswith(b) and not f.startswith(b + "-") and len(f) > len(b):
        return f"{b}-{f[len(b):]}"
    return f


def _hkey(h): return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


HEADERS = {
    "block":  ("block", "tower", "wing", "building"),
    "flat":   ("flat", "flatno", "flatnumber", "unit", "unitno", "house", "houseno", "apartment"),
    "name":   ("ownername", "name", "resident", "residentname", "owner", "fullname"),
    "phone":  ("mobile", "phone", "whatsapp", "contact", "mobileno", "phonenumber", "mobilenumber"),
    "model":  ("model", "vehiclemodel", "carmodel", "make"),
    "color":  ("color", "colour", "vehiclecolor"),
}


def _map_headers(headers):
    """-> (col_index_by_field, [plate column indexes])"""
    keys = [_hkey(h) for h in headers]
    idx, plates = {}, []
    for i, k in enumerate(keys):
        if not k:
            continue
        if any(t in k for t in ("plate", "vehicle", "reg", "carno", "vehno")) and "model" not in k \
                and "color" not in k and "colour" not in k and "type" not in k:
            plates.append(i); continue
        for field, aliases in HEADERS.items():
            if field not in idx and k in aliases:
                idx[field] = i
    return idx, plates


# ════════════════════════════════════════════════════════════════════
# Parse
# ════════════════════════════════════════════════════════════════════

def _read_rows(file_storage):
    name = (file_storage.filename or "").lower()
    data = file_storage.read()
    if name.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def parse_sheet(rows):
    """-> {flats: {flat_no: {...}}, vehicles: [..], issues: [..]}"""
    rows = [r for r in rows if r and any(c not in (None, "") for c in r)]
    if not rows:
        return {"error": "The sheet is empty"}
    idx, plate_cols = _map_headers(rows[0])
    if "flat" not in idx:
        return {"error": "Couldn't find a 'flat' column. Use the template or name the column 'flat'."}
    flats, vehicles, issues = {}, [], []
    seen_plates = {}
    for n, r in enumerate(rows[1:], start=2):
        g = lambda f: (str(r[idx[f]]).strip() if f in idx and idx[f] < len(r) and r[idx[f]] is not None else "")
        flat_no = _norm_flat(g("block"), g("flat"))
        if not flat_no:
            issues.append({"row": n, "level": "error", "text": "No flat number"}); continue
        name = g("name")
        phone_raw = g("phone")
        phone = _norm_phone(phone_raw)
        if phone_raw and not phone:
            issues.append({"row": n, "level": "error", "flat": flat_no,
                           "text": f"Mobile '{phone_raw}' isn't a 10-digit Indian number"})
        f = flats.setdefault(flat_no, {"flat_no": flat_no, "owner_name": "", "whatsapp": "", "row": n})
        if name and not f["owner_name"]:
            f["owner_name"] = name
        if phone and not f["whatsapp"]:
            f["whatsapp"] = phone
        elif phone and f["whatsapp"] and phone != f["whatsapp"]:
            issues.append({"row": n, "level": "warn", "flat": flat_no,
                           "text": f"Second mobile for {flat_no} ignored ({phone}); app login uses {f['whatsapp']}"})
        for ci in plate_cols:
            raw = r[ci] if ci < len(r) else None
            plate = _norm_plate(raw)
            if not plate:
                continue
            if not PLATE_RE.match(plate):
                issues.append({"row": n, "level": "warn", "flat": flat_no,
                               "text": f"'{raw}' doesn't look like an Indian plate — imported as typed"})
            if plate in seen_plates and seen_plates[plate] != flat_no:
                issues.append({"row": n, "level": "error", "flat": flat_no,
                               "text": f"{plate} is also listed under {seen_plates[plate]} — kept the first"})
                continue
            seen_plates[plate] = flat_no
            vehicles.append({"plate": plate, "flat_no": flat_no, "name": name or f["owner_name"],
                             "phone": phone or f["whatsapp"], "model": g("model"), "color": g("color")})
    no_phone = [k for k, v in flats.items() if not v["whatsapp"]]
    if no_phone:
        issues.append({"row": 0, "level": "warn", "flat": "",
                       "text": f"{len(no_phone)} flat(s) have no mobile — they can't log in to the app: "
                               + ", ".join(no_phone[:8]) + (" …" if len(no_phone) > 8 else "")})
    return {"flats": flats, "vehicles": vehicles, "issues": issues}


# ════════════════════════════════════════════════════════════════════
# Write
# ════════════════════════════════════════════════════════════════════

def _write_flats(flats: dict) -> tuple:
    """Return (written, how, skipped_no_phone). Uses flat_directory.set_flat —
    the same function the CLI uses — so validation rules stay in one place.
    Flats without a mobile are skipped here (they can't log in anyway) but
    their vehicles are still imported."""
    try:
        import flat_directory as fd
    except Exception as e:
        return 0, f"flat_directory unavailable: {e}", 0
    if callable(getattr(fd, "set_flat", None)):
        try:
            fd.init_flats()
        except Exception:
            pass
        n, skipped = 0, 0
        for v in flats.values():
            if not v["whatsapp"]:
                skipped += 1
                continue
            try:
                ok = fd.set_flat(v["flat_no"], v["owner_name"] or "Resident", v["whatsapp"])
            except Exception as e:
                logger.warning(f"[IMPORT] set_flat {v['flat_no']}: {e}")
                ok = False
            n += 1 if ok else 0
        return n, "flat_directory.set_flat", skipped
    # Fallback for a different flat_directory build: its own table, detected by name
    try:
        con = sqlite3.connect(_DB_PATH)
        tables = [r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")]
        for t in tables:
            if "flat" not in t.lower():
                continue          # only a flats table — never visitors/residents
            cols = [c[1] for c in con.execute(f"PRAGMA table_info({t})")]
            lc = {c.lower(): c for c in cols}
            flat_col = next((lc[k] for k in ("flat_no", "flat", "flat_number", "unit") if k in lc), None)
            ph_col = next((lc[k] for k in ("whatsapp", "phone", "mobile", "owner_phone") if k in lc), None)
            nm_col = next((lc[k] for k in ("owner_name", "name", "resident_name") if k in lc), None)
            if flat_col and ph_col:
                n, skipped = 0, 0
                for v in flats.values():
                    if not v["whatsapp"]:
                        skipped += 1; continue
                    ex = con.execute(f"SELECT 1 FROM {t} WHERE UPPER({flat_col})=?", (v["flat_no"],)).fetchone()
                    if ex:
                        sets, vals = [f"{ph_col}=?"], [v["whatsapp"]]
                        if nm_col and v["owner_name"]:
                            sets.append(f"{nm_col}=?"); vals.append(v["owner_name"])
                        con.execute(f"UPDATE {t} SET {', '.join(sets)} WHERE UPPER({flat_col})=?", vals + [v["flat_no"]])
                    else:
                        cols_i, vals = [flat_col, ph_col], [v["flat_no"], v["whatsapp"]]
                        if nm_col:
                            cols_i.append(nm_col); vals.append(v["owner_name"] or "Resident")
                        con.execute(f"INSERT INTO {t} ({', '.join(cols_i)}) VALUES ({', '.join('?' * len(vals))})", vals)
                    n += 1
                con.commit(); con.close()
                return n, f"table {t}", skipped
        con.close()
        return 0, "no flats table found", 0
    except Exception as e:
        return 0, f"flat table write failed: {e}", 0


def _write_vehicles(vehicles: list) -> int:
    from resident_db import db as rdb, Resident
    n = 0
    for v in vehicles:
        fl, bl = v["flat_no"], ""
        if "-" in fl:
            bl, fl = fl.split("-", 1)
        rdb.add(Resident(plate_number=v["plate"], resident_name=v["name"] or "Resident",
                         flat_number=fl, block=bl, phone=v["phone"], vehicle_model=v["model"],
                         vehicle_color=v["color"], status="KNOWN", notes="bulk import"))
        n += 1
    return n


# ════════════════════════════════════════════════════════════════════
# Routes
# ════════════════════════════════════════════════════════════════════

@resident_import_bp.route("/api/admin/residents/import", methods=["POST"])
def import_residents():
    if "file" not in request.files:
        return jsonify({"success": False, "message": "Upload the Excel or CSV as 'file'"}), 400
    dry = (request.form.get("dry_run", "1") != "0")
    f = request.files["file"]
    try:
        rows = _read_rows(f)
    except Exception as e:
        return jsonify({"success": False, "message": f"Couldn't read the file: {e}"}), 400
    parsed = parse_sheet(rows)
    if "error" in parsed:
        return jsonify({"success": False, "message": parsed["error"]}), 400
    flats, vehicles, issues = parsed["flats"], parsed["vehicles"], parsed["issues"]
    out = {
        "success": True, "dry_run": dry,
        "flats": len(flats), "flats_with_phone": sum(1 for v in flats.values() if v["whatsapp"]),
        "vehicles": len(vehicles),
        "errors": sum(1 for i in issues if i["level"] == "error"),
        "warnings": sum(1 for i in issues if i["level"] == "warn"),
        "issues": issues[:60],
        "preview": [{"flat_no": v["flat_no"], "owner_name": v["owner_name"], "whatsapp": v["whatsapp"],
                     "vehicles": [x["plate"] for x in vehicles if x["flat_no"] == v["flat_no"]]}
                    for v in list(flats.values())[:8]],
    }
    if dry:
        return jsonify(out)
    fw, how, skipped = _write_flats(flats)
    vw = _write_vehicles(vehicles)
    out.update({"flats_written": fw, "flats_how": how, "flats_skipped_no_phone": skipped,
                "vehicles_written": vw})
    try:
        con = sqlite3.connect(_DB_PATH)
        con.execute("INSERT INTO resident_imports (filename, flats, vehicles, issues, imported_at, by_user) "
                    "VALUES (?,?,?,?,?,?)",
                    (f.filename, fw, vw, len(issues), datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     (getattr(request, "auth_user", None) or {}).get("username") or "admin"))
        con.commit(); con.close()
    except sqlite3.Error:
        pass
    logger.info(f"[IMPORT] {f.filename}: {fw} flats ({how}), {vw} vehicles, {len(issues)} issues")
    return jsonify(out)


def _all_flats():
    try:
        from resident_app import _all_directory_flats
        return _all_directory_flats()
    except Exception:
        return []


@resident_import_bp.route("/api/admin/residents/summary")
def residents_summary():
    flats = _all_flats()
    phones = {_norm_phone(f.get("whatsapp") or f.get("phone") or "") for f in flats}
    phones.discard("")
    try:
        from resident_db import db as rdb
        vehicles = rdb.count()
    except Exception:
        vehicles = 0
    con = sqlite3.connect(_DB_PATH); con.row_factory = sqlite3.Row
    def q(sql, *a):
        try: return con.execute(sql, a).fetchone()[0]
        except sqlite3.Error: return 0
    out = {
        "success": True,
        "flats": len(flats), "flats_with_phone": len(phones), "vehicles": vehicles,
        "app_users": q("SELECT COUNT(*) FROM resident_logins"),
        "invited": q("SELECT COUNT(*) FROM resident_invites WHERE status='sent'"),
        "passes_30d": q("SELECT COUNT(*) FROM gate_passes WHERE created_at >= date('now','-30 days')"),
        "arrivals_30d": q("SELECT COUNT(*) FROM arrival_requests WHERE created_at >= date('now','-30 days')"),
        "sos_30d": q("SELECT COUNT(*) FROM resident_sos WHERE created_at >= date('now','-30 days')"),
        "last_import": None,
        "invite_job": dict(_invite_job),
    }
    try:
        r = con.execute("SELECT * FROM resident_imports ORDER BY id DESC LIMIT 1").fetchone()
        if r: out["last_import"] = dict(r)
    except sqlite3.Error:
        pass
    con.close()
    return jsonify(out)


@resident_import_bp.route("/api/admin/residents/template")
def residents_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Residents"
    ws.append(["Block", "Flat", "Owner Name", "Mobile", "Vehicle 1", "Vehicle 2", "Vehicle 3", "Model"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="0E7490")
    ws.append(["B", "302", "Akash Singh", "9876543210", "PB08EY5332", "", "", "Hyundai i20"])
    ws.append(["A", "105", "Gurpreet Kaur", "9810012345", "PB10AB2025", "PB10CD7788", "", "Swift / City"])
    for col, w in zip("ABCDEFGH", (8, 8, 24, 14, 14, 14, 14, 20)):
        ws.column_dimensions[col].width = w
    b = io.BytesIO(); wb.save(b); b.seek(0)
    return send_file(b, as_attachment=True, download_name="DefenderOcta_Residents_Template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _welcome_text(site, flat, name):
    return (f"👋 *Defender Octa* — {site}\n\n"
            f"Hi {name or 'there'}, your society has switched on Defender Octa, the AI gate "
            f"security system. Flat {flat} now has its own app:\n\n"
            f"🔗 {_app_link()}\n\n"
            f"See who came to your flat, pre-approve guests with a gate pass, approve "
            f"visitors from your phone, and raise an SOS that reaches the guard in seconds.\n"
            f"Log in with this WhatsApp number — no password.\n"
            f"_Reply STOP to opt out._")


def _app_link():
    return os.environ.get("OCTA_PUBLIC_URL", "").rstrip("/") + "/resident" \
        if os.environ.get("OCTA_PUBLIC_URL") else "https://agi.snguardiangrid.com/resident"


def _site_name():
    try:
        from resident_app import _site_name as sn
        return sn()
    except Exception:
        return "Your society"


@resident_import_bp.route("/api/admin/residents/invite", methods=["POST"])
def invite_residents():
    if _invite_job["running"]:
        return jsonify({"success": False, "message": "An invite run is already in progress"}), 409
    data = request.get_json(silent=True) or {}
    only = {_norm_phone(p) for p in (data.get("phones") or [])} - {""}
    sample = int(data.get("sample") or 0)
    custom = (data.get("message") or "").strip()
    site = _site_name()
    targets = []
    for f in _all_flats():
        ph = _norm_phone(f.get("whatsapp") or f.get("phone") or "")
        if not ph or (only and ph not in only):
            continue
        targets.append((ph, f.get("flat_no") or f.get("flat") or "", f.get("owner_name") or f.get("name") or ""))
    if sample:
        targets = targets[:sample]
    if not targets:
        return jsonify({"success": False, "message": "No mobiles to invite — import residents first"}), 400

    def _run():
        from resident_app import _send_wa
        _invite_job.update(running=True, total=len(targets), sent=0, failed=0,
                           started=datetime.now().strftime("%H:%M:%S"), finished=None, last_error="")
        con = sqlite3.connect(_DB_PATH)
        for ph, flat, name in targets:
            body = custom.replace("{flat}", flat).replace("{name}", name or "there").replace("{link}", _app_link()) \
                if custom else _welcome_text(site, flat, name)
            r = _send_wa(ph, body)
            ok = bool(r.get("success"))
            with _lock:
                _invite_job["sent" if ok else "failed"] += 1
                if not ok: _invite_job["last_error"] = str(r.get("error", ""))[:120]
            con.execute("INSERT OR REPLACE INTO resident_invites (phone, flat_no, sent_at, status, detail) "
                        "VALUES (?,?,?,?,?)", (ph, flat, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                              "sent" if ok else "failed", str(r.get("error") or r.get("sid") or "")[:120]))
            con.commit()
            time.sleep(0.4)          # gentle on Twilio's rate limit
        con.close()
        _invite_job.update(running=False, finished=datetime.now().strftime("%H:%M:%S"))

    threading.Thread(target=_run, daemon=True, name="resident-invite").start()
    return jsonify({"success": True, "queued": len(targets),
                    "message": f"Inviting {len(targets)} resident{'s' if len(targets) != 1 else ''} — "
                               f"watch the counter"})


@resident_import_bp.route("/api/admin/residents/invite/status")
def invite_status():
    return jsonify({"success": True, **_invite_job})
