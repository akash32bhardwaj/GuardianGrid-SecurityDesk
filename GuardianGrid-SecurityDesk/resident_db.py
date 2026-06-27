import os
import json
import csv
import logging
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional

# Import DB if available
from database.db import get_connection, DATABASE_URL, init_db

logger = logging.getLogger(__name__)

# ── Try importing openpyxl for Excel support ─────────────────────
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Data paths ───────────────────────────────────────────────────
DB_FILE      = Path("data/residents.json")
DB_FILE.parent.mkdir(parents=True, exist_ok=True)


# ── Resident dataclass ───────────────────────────────────────────
@dataclass
class Resident:
    plate_number:   str
    resident_name:  str
    flat_number:    str
    block:          str        = ""
    phone:          str        = ""
    vehicle_type:   str        = "Car"
    vehicle_model:  str        = ""
    vehicle_color:  str        = ""
    status:         str        = "KNOWN"   # KNOWN / BLACKLISTED / VISITOR
    notes:          str        = ""
    added_on:       str        = field(
        default_factory=lambda: datetime.now().strftime("%Y-%m-%d")
    )

    def to_dict(self):
        return asdict(self)

    @property
    def display_name(self):
        parts = [self.resident_name]
        if self.flat_number:
            parts.append(f"Flat {self.flat_number}")
        if self.block:
            parts.append(f"Block {self.block}")
        return " · ".join(parts)


# ── Database class ───────────────────────────────────────────────
class ResidentDatabase:
    def __init__(self):
        self._db: dict = {}    # plate_number -> Resident
        self.use_pg = bool(DATABASE_URL)
        if self.use_pg:
            init_db()
        self._load()

    # ── Persistence ──────────────────────────────────────────
    def _load(self):
        if self.use_pg:
            self._db.clear()
            try:
                conn = get_connection()
                with conn.cursor() as cur:
                    cur.execute("SELECT plate_number, resident_name, flat_number, block, phone, vehicle_type, vehicle_model, vehicle_color, status, notes, added_on FROM resident_vehicles")
                    for row in cur.fetchall():
                        r = Resident(*row)
                        self._db[r.plate_number] = r
                conn.close()
                logger.info(f"Loaded {len(self._db)} residents from PostgreSQL.")
            except Exception as e:
                logger.error(f"Could not load resident DB from PG: {e}")
        else:
            if DB_FILE.exists():
                try:
                    with open(DB_FILE, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    for plate, record in data.items():
                        self._db[plate] = Resident(**record)
                    logger.info(f"Loaded {len(self._db)} residents from JSON database.")
                except Exception as e:
                    logger.error(f"Could not load resident DB from JSON: {e}")

    def _save(self):
        if not self.use_pg:
            with open(DB_FILE, "w", encoding="utf-8") as f:
                json.dump(
                    {plate: r.to_dict() for plate, r in self._db.items()},
                    f, indent=2, ensure_ascii=False
                )

    def _pg_upsert(self, resident: Resident):
        if not self.use_pg:
            return
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO resident_vehicles 
                    (plate_number, resident_name, flat_number, block, phone, vehicle_type, vehicle_model, vehicle_color, status, notes, added_on)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (plate_number) DO UPDATE SET
                    resident_name=EXCLUDED.resident_name,
                    flat_number=EXCLUDED.flat_number,
                    block=EXCLUDED.block,
                    phone=EXCLUDED.phone,
                    vehicle_type=EXCLUDED.vehicle_type,
                    vehicle_model=EXCLUDED.vehicle_model,
                    vehicle_color=EXCLUDED.vehicle_color,
                    status=EXCLUDED.status,
                    notes=EXCLUDED.notes,
                    added_on=EXCLUDED.added_on
                """, (
                    resident.plate_number, resident.resident_name, resident.flat_number,
                    resident.block, resident.phone, resident.vehicle_type, resident.vehicle_model,
                    resident.vehicle_color, resident.status, resident.notes, resident.added_on
                ))
            conn.commit()
        except Exception as e:
            logger.error(f"Error saving to PG: {e}")
        finally:
            conn.close()

    def _pg_delete(self, plate: str):
        if not self.use_pg:
            return
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM resident_vehicles WHERE plate_number = %s", (plate,))
            conn.commit()
        except Exception as e:
            logger.error(f"Error deleting from PG: {e}")
        finally:
            conn.close()

    # ── Core operations ───────────────────────────────────────
    def lookup(self, plate: str) -> Optional[Resident]:
        clean = plate.upper().replace(" ", "")
        if clean in self._db:
            return self._db[clean]
        for stored_plate, resident in self._db.items():
            if clean in stored_plate or stored_plate in clean:
                return resident
        return None

    def fuzzy_lookup(self, plate: str, max_distance: int = 1) -> Optional[Resident]:
        clean = plate.upper().replace(" ", "")
        if clean in self._db:
            return self._db[clean]

        candidates = []
        for stored_plate, resident in self._db.items():
            if len(stored_plate) != len(clean):
                continue
            dist = sum(1 for a, b in zip(clean, stored_plate) if a != b)
            if dist <= max_distance:
                candidates.append((dist, resident))

        if len(candidates) == 1:
            return candidates[0][1]
        if len(candidates) > 1:
            candidates.sort(key=lambda c: c[0])
            if candidates[0][0] < candidates[1][0]:
                return candidates[0][1]
        return None

    def add(self, resident: Resident):
        clean = resident.plate_number.upper().replace(" ", "")
        resident.plate_number = clean
        self._db[clean] = resident
        self._pg_upsert(resident)
        self._save()

    def remove(self, plate: str) -> bool:
        clean = plate.upper().replace(" ", "")
        if clean in self._db:
            del self._db[clean]
            self._pg_delete(clean)
            self._save()
            return True
        return False

    def get_all(self) -> list:
        return [r.to_dict() for r in self._db.values()]

    def count(self) -> int:
        return len(self._db)

    def blacklist(self, plate: str, reason: str = ""):
        r = self.lookup(plate)
        if r:
            r.status = "BLACKLISTED"
            r.notes  = reason
            self.add(r)

    def import_from_excel(self, filepath: str) -> dict:
        if not EXCEL_AVAILABLE:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(filepath)
        if not path.exists():
            return {"error": f"File not found: {filepath}"}

        try:
            wb   = openpyxl.load_workbook(path)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return {"error": f"Cannot read Excel file: {e}"}

        if not rows:
            return {"error": "Excel file is empty"}

        headers = [str(h).lower().strip() if h else "" for h in rows[0]]

        col_map = {
            "plate_number":  ["plate", "plate_number", "vehicle_number", "registration", "reg_no", "number_plate"],
            "resident_name": ["name", "resident_name", "owner", "resident", "full_name", "owner_name"],
            "flat_number":   ["flat", "flat_number", "flat_no", "unit", "apartment", "house_no", "unit_no"],
            "block":         ["block", "tower", "wing", "sector"],
            "phone":         ["phone", "mobile", "contact", "phone_number", "mobile_number"],
            "vehicle_type":  ["type", "vehicle_type", "vehicle"],
            "vehicle_model": ["model", "vehicle_model", "car_model"],
            "vehicle_color": ["color", "colour", "vehicle_color"],
            "notes":         ["notes", "remarks", "comment"],
        }

        col_idx = {}
        for field_name, aliases in col_map.items():
            for alias in aliases:
                if alias in headers:
                    col_idx[field_name] = headers.index(alias)
                    break

        if "plate_number" not in col_idx:
            return {"error": "Could not find plate number column. Please name it 'plate' or 'plate_number'."}

        imported = 0
        skipped  = 0
        errors   = []

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                plate = str(row[col_idx["plate_number"]] or "").strip()
                if not plate or plate.lower() == "none":
                    skipped += 1
                    continue

                def get_col(field, default=""):
                    idx = col_idx.get(field)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val is not None else default
                    return default

                resident = Resident(
                    plate_number  = plate.upper().replace(" ", ""),
                    resident_name = get_col("resident_name", "Unknown"),
                    flat_number   = get_col("flat_number"),
                    block         = get_col("block"),
                    phone         = get_col("phone"),
                    vehicle_type  = get_col("vehicle_type", "Car"),
                    vehicle_model = get_col("vehicle_model"),
                    vehicle_color = get_col("vehicle_color"),
                    notes         = get_col("notes"),
                    status        = "KNOWN",
                )
                self.add(resident)
                imported += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {e}")
                skipped += 1

        return {
            "imported": imported,
            "skipped":  skipped,
            "total":    len(rows) - 1,
            "errors":   errors,
        }

    def export_to_excel(self, filepath: str = "data/residents_export.xlsx"):
        if not EXCEL_AVAILABLE:
            return {"error": "openpyxl not installed"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Residents"

        headers = [
            "Plate Number", "Resident Name", "Flat Number", "Block",
            "Phone", "Vehicle Type", "Vehicle Model",
            "Vehicle Color", "Status", "Notes", "Added On"
        ]
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")

        for r in self._db.values():
            ws.append([
                r.plate_number, r.resident_name, r.flat_number,
                r.block, r.phone, r.vehicle_type, r.vehicle_model,
                r.vehicle_color, r.status, r.notes, r.added_on
            ])

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        wb.save(filepath)
        return {"exported": len(self._db), "file": filepath}

    def import_from_csv(self, filepath: str) -> dict:
        path = Path(filepath)
        if not path.exists():
            return {"error": f"File not found: {filepath}"}

        imported = 0
        skipped  = 0

        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                plate = (row.get("plate_number") or
                         row.get("plate") or "").strip().upper()
                if not plate:
                    skipped += 1
                    continue
                resident = Resident(
                    plate_number  = plate.replace(" ", ""),
                    resident_name = row.get("resident_name", "Unknown"),
                    flat_number   = row.get("flat_number", ""),
                    block         = row.get("block", ""),
                    phone         = row.get("phone", ""),
                    vehicle_type  = row.get("vehicle_type", "Car"),
                    vehicle_model = row.get("vehicle_model", ""),
                    vehicle_color = row.get("vehicle_color", ""),
                    notes         = row.get("notes", ""),
                )
                self.add(resident)
                imported += 1

        return {"imported": imported, "skipped": skipped}

# ── Global instance ───────────────────────────────────────────────
db = ResidentDatabase()

# ── Standalone test ───────────────────────────────────────────────
if __name__ == "__main__":
    print("Testing Resident Database...")
    print(f"Total residents: {db.count()}")
